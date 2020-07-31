from openpyxl import Workbook, load_workbook

wb = load_workbook(filename = 'data.xlsx')

outputWb = Workbook()

startColumn = 3

sheetNames = wb.sheetnames


for sheet in sheetNames:
  lastRow = 5
  used = []
  data = {}
  workSheet = wb[sheet]
  print(sheet)
  sheetId = sheetNames.index(sheet)
  while True:
    rowId = workSheet.cell(column=2, row=lastRow).value
    rowName = workSheet.cell(column=1, row=lastRow).value
    if (rowId == None): #Если инн пустой
      break
    
    for columnOffset in range(0, 20):
      column = startColumn + columnOffset
      cellValue = workSheet.cell(column=column, row=lastRow).value
      if (cellValue == None): # если ячейка пустая
        rowOffset = 0
        while True:
          secondRowId = workSheet.cell(column=2, row=lastRow + rowOffset).value
          if (secondRowId == None or secondRowId != rowId):
            break
          secondCellValue = workSheet.cell(column=column, row=lastRow + rowOffset).value
          if ((secondRowId == rowId) and (secondCellValue != None)):
            if secondRowId in data.keys():
              data[secondRowId]['columns'].append({
                'offset': column,
                'value': secondCellValue
              })
            else:
              data[secondRowId] = {
                'name': rowName,
                'columns': [{
                    'offset': column,
                    'value': secondCellValue
                  }]
              }
          rowOffset += 1
      else:
        if rowId in data.keys():
          data[rowId]['columns'].append(
              {
                'offset': column,
                'value': cellValue
              }
            )
        else:
          data[rowId] = {
            'name': rowName,
            'columns': [{
                'offset': column,
                'value': cellValue
              }]
          }
    lastRow += 1

  rowNumber = 1
  outputSheet = outputWb.create_sheet(title=sheet)

  for (key, value) in data.items():
    outputSheet.cell(column=1, row=rowNumber, value=value['name'])
    outputSheet.cell(column=2, row=rowNumber, value=key)
    for column in value['columns']:
      outputSheet.cell(column=column["offset"], row=rowNumber, value=column["value"])
    rowNumber += 1    
  outputWb.save(filename = 'out.xlsx')

